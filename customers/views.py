# =============================
# IMPORTS
# =============================

import json
import os
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from decimal import Decimal
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook, load_workbook
from accounts.decorators import token_required
from cooking_items.forms import CookingItemForm
from ingredients.forms import IngredientForm
from dishes.forms import DishForm
from .models import (
    Customer,
    CustomerCookingItem,
    CustomerDish,
    CustomerIngredient,
)

from .forms import CustomerForm, CustomerDishFormSet, CustomerCookingItemFormSet

from dishes.models import Dish, DishIngredient
from cooking_items.models import CookingItem
from ingredients.models import Ingredient


# =============================
# BASIC VIEWS
# =============================
def add_watermark(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 60)
    canvas.setFillColorRGB(0.92, 0.92, 0.92)
    canvas.rotate(45)
    canvas.drawCentredString(400, 100, "INVOICE")
    canvas.restoreState()


def add_page_number(canvas, doc):
    page_num = canvas.getPageNumber()
    text = f"Page {page_num}"
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(580, 15, text)


def home(request):
    return render(request, 'home.html')


def logout_view(request):
    logout(request)
    return redirect('home')


@login_required
def dashboard(request):
    return render(request, 'dashboard.html')


# =============================
# CUSTOMER ADD / EDIT
# =============================
@token_required
@login_required  
def customer_add_edit(request, pk=None):

    customer = None

    if pk:
        customer = get_object_or_404(Customer, pk=pk)

    if request.method == "POST":
        form = CustomerForm(request.POST, request.FILES, instance=customer)

        if form.is_valid():
            customer = form.save()

            # Save selected dishes
            selected_dishes = request.POST.getlist("selected_dishes")
            customer.dishes.set(selected_dishes)

            
            # Save selected Ingredients
            # ------------------------------
            CustomerIngredient.objects.filter(customer=customer).delete()

            selected_ingredients = request.POST.getlist("selected_ingredients")

            for ing_id in selected_ingredients:
                CustomerIngredient.objects.create(
                    customer=customer,
                    ingredients_id=ing_id,
                    quantity=1
                )

            # Save selected cooking items
            selected_cooking = request.POST.getlist("selected_cooking")
            customer.cooking_items.set(selected_cooking)

            return redirect("customer_list")
    else:
        form = CustomerForm(instance=customer)

    # ----------------------
    # Prepare Dish JSON
    # ----------------------

    dish_json = {}
    for dish in Dish.objects.all():

        ingredients = DishIngredient.objects.filter(dish=dish)

        dish_json[dish.id] = {
            "name": dish.name_en,
            "ingredients": [
                {
                    "name": di.ingredient.name_en,
                    "quantity": di.quantity,
                    "unit": di.unit,
                    "price": float(di.price),
                }
                for di in ingredients
            ],
        }

    # ----------------------
    # Prepare Ingredient JSON
    # ----------------------

    ingredients_json = {}
    for item in Ingredient.objects.all():
        ingredients_json[item.id] = {
            "name": item.name_en,
            "price": float(item.price),  
        }

    # ----------------------
    # Prepare Cooking JSON
    # ----------------------

    cooking_json = {}
    for item in CookingItem.objects.all():
        cooking_json[item.id] = {
            "name": item.name_en,
            "price": float(item.cost),  # IMPORTANT: use cost not price
        }

    # ----------------------
    # Load Selected Data For Edit
    # ----------------------

    selected_dishes = []
    selected_cooking = []
    selected_ingredients = []

    if customer:
        selected_dishes = list(customer.dishes.values_list("id", flat=True))
        selected_cooking = list(customer.cooking_items.values_list("id", flat=True))
        selected_ingredients = list(
            CustomerIngredient.objects.filter(customer=customer)
            .values_list("ingredients_id", flat=True)

        )

    context = {
    "form": form,
    "form_title": "Edit Customer" if customer else "Add Customer",
    "dish_dropdown": {d.id: d.name_en for d in Dish.objects.all()},
    "ingredient_dropdown": {i.id: i.name_en for i in Ingredient.objects.all()},
    "cooking_items": CookingItem.objects.all(),
    "dish_json": json.dumps(dish_json),
    "cooking_json": json.dumps(cooking_json),
    "ingredient_json": json.dumps(ingredients_json),
    "selected_dishes": json.dumps(selected_dishes),
    "selected_cooking": json.dumps(selected_cooking),
    "selected_ingredients": json.dumps(selected_ingredients),
    }

    return render(request, "customers/add_customer.html", context)

# =============================
# ADD DISH / COOKING / INGREDIENT ITEM
# =============================
@token_required
@login_required
def dish_add(request):
    if request.method == 'POST':
        form = DishForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_add')
    else:
        form = DishForm()

    return render(request, 'customers/dish_form.html', {
        'form_title': 'Add Dish',
        'form': form
    })

@token_required
@login_required
def cooking_item_add(request):
    if request.method == 'POST':
        form = CookingItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_add')
    else:
        form = CookingItemForm()

    return render(request, 'customers/cooking_item_form.html', {
        'form_title': 'Add Cooking Item',
        'form': form
    })

@token_required
@login_required
def ingredient_item_add(request, customer_id):
    if request.method == 'POST':
        form = IngredientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_add')
    else:
        form = CookingItemForm()

    return render(request, 'customers/ingredient_form.html', {
        'form_title': 'Add Ingredient Item',
        'form': form
    })
# =============================
# CUSTOMER LIST
# =============================

@login_required
def customer_list(request):
    query = request.GET.get("q")

    if query:
        customers = Customer.objects.filter(name__icontains=query)
    else:
        customers = Customer.objects.all()

    return render(request, "customers/customer_list.html", {
        "customers": customers
    })


# =============================
# DELETE CUSTOMER
# =============================
@token_required
@login_required
def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    customer.delete()
    messages.success(request, "Customer deleted successfully!")
    return redirect("customer_list")


# =============================
# EXPORT CUSTOMERS (EXCEL)
# =============================
@token_required
@login_required
def export_customers(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = ["Name", "Phone"]
    ws.append(headers)

    customers = Customer.objects.all()

    for customer in customers:
        ws.append([
            customer.name,
            customer.phone
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'

    wb.save(response)
    return response


# =============================
# IMPORT CUSTOMERS (EXCEL)
# =============================
@token_required
@login_required
def import_customers(request):
    if request.method == "POST":
        excel_file = request.FILES.get("file")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("customer_list")

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, phone = row

            if name:
                Customer.objects.create(
                    name=name,
                    phone=phone,
                    created_by=request.user
                )

        messages.success(request, "Customers imported successfully!")
        return redirect("customer_list")

    return render(request, "customers/import_customers.html")


# =============================
# AJAX: GET DISH INGREDIENTS
# =============================
@token_required
@login_required
def get_dish_ingredients(request, dish_id):
    dish = get_object_or_404(Dish, id=dish_id)

    ingredients = []
    for ingredient in dish.ingredients.all():
        ingredients.append({
            'name': ingredient.name
        })

    return JsonResponse({'ingredients': ingredients})

# =============================
# GENERATE INGREDIENT LIST (PDF)
# =============================
@token_required
@login_required
def generate_ingredient_list(request, customer_id):

    customer = get_object_or_404(Customer, id=customer_id)
    owner = request.user

    dishes = customer.dishes.all()
    cooking_items = customer.cooking_items.all()
    customer_ingredients = customer.customeringredient_set.all()

    persons = Decimal(customer.num_person or 1)

    ingredient_summary = {}
    cooking_summary = {}
    Ingredient_count = 0
    cooking_items_count = 0
    total_ingredient_cost = Decimal("0.00")
    total_cooking_cost = Decimal("0.00")

    # =====================================================
    # 1️⃣ DISH INGREDIENT CALCULATION
    # =====================================================
    for dish in dishes:

        dish_ingredients = DishIngredient.objects.filter(dish=dish)

        for di in dish_ingredients:

            quantity = Decimal(di.quantity)
            price = Decimal(di.price)

            total_qty = quantity * persons
            total_price = total_qty * price

            name = di.ingredient.name_en

            if name in ingredient_summary:
                ingredient_summary[name]["qty"] += total_qty
                ingredient_summary[name]["total"] += total_price
            else:
                Ingredient_count += 1
                ingredient_summary[name] = {
                    "qty": total_qty,
                    "unit": di.unit,
                    "price": price,
                    "total": total_price,
                }

            total_ingredient_cost += total_price

    # =====================================================
    # 2️⃣ CUSTOMER EXTRA INGREDIENTS
    # =====================================================
    for ci in customer_ingredients:

        ing = ci.ingredients

        quantity = Decimal(ci.quantity)
        price = Decimal(ing.price)

        total_qty = quantity * persons
        item_total = total_qty * price

        name = ing.name_en

        if name in ingredient_summary:
            ingredient_summary[name]["qty"] += total_qty
            ingredient_summary[name]["total"] += item_total
        else:
            Ingredient_count += 1
            ingredient_summary[name] = {
                "qty": total_qty,
                "unit": ing.unit,
                "price": price,
                "total": item_total,
            }

        total_ingredient_cost += item_total

    # =====================================================
    # 3️⃣ COOKING ITEMS CALCULATION
    # =====================================================
    for item in cooking_items:

        price = Decimal(item.cost)
        total_qty = Decimal("1.00")
        item_total = price

        name = item.name_en

        cooking_summary[name] = {
            "qty": total_qty,
            "price": price,
            "total": item_total,
        }

        cooking_items_count += 1
        total_cooking_cost += item_total

    # =====================================================
    # GRAND TOTAL
    # =====================================================
    grand_total = total_ingredient_cost + total_cooking_cost

    context = {
        "customer": customer,
        "owner": owner,
        "dishes": dishes,
        "ingredient_summary": ingredient_summary,
        "cooking_summary": cooking_summary,
        "total_ingredient_cost": total_ingredient_cost,
        "total_cooking_cost": total_cooking_cost,
        "grand_total": grand_total,
        "cooking_items_count": cooking_items_count,
        "Ingredient_count": Ingredient_count,
    }

    return render(request, "customers/generate_summary.html", context)

# =============================
# DOWNLOAD CUSTOMER PDF
# =============================

@token_required
@login_required

# =============================
# EXPORT CUSTOMER PDF
# =============================
@token_required
@login_required
# def export_customer_pdf(request, customer_id):

#     customer = get_object_or_404(Customer, id=customer_id)

#     # 🔥 Use customer.created_by instead of request.user
#     customer = get_object_or_404(Customer, id=customer_id)
#     owner = request.user

#     dishes = customer.dishes.all()
#     cooking_items = customer.cooking_items.all()
#     customer_ingredients = customer.customeringredient_set.all()

#     persons = customer.num_person or 1   # ✅ your correct field
#     Ingredient_count = 0
#     cooking_items_count = 0
#     ingredient_summary = {}
#     cooking_summary = {}

#     total_ingredient_cost = Decimal("0.00")
#     total_cooking_cost = Decimal("0.00")

#     # -----------------------------------
#     # Calculate Ingredients
#     # -----------------------------------
#     # =====================================================
#     # 1️⃣ DISH INGREDIENT CALCULATION
#     # =====================================================
#     for dish in dishes:

#         dish_ingredients = DishIngredient.objects.filter(dish=dish)

#         for di in dish_ingredients:

#             total_qty = di.quantity * persons
#             total_price = Decimal(total_qty) * Decimal(di.price)

#             name = di.ingredient.name_en

#             if name in ingredient_summary:
#                 ingredient_summary[name]["qty"] += total_qty
#                 ingredient_summary[name]["total"] += total_price
#             else:
#                 Ingredient_count += 1
#                 ingredient_summary[name] = {
#                     "qty": total_qty,
#                     "unit": di.unit,
#                     "price": di.price,
#                     "total": total_price,
#                 }

#             total_ingredient_cost += total_price

#     # =====================================================
#     # 2️⃣ CUSTOMER EXTRA INGREDIENTS
#     # =====================================================
#     for ci in customer_ingredients:

#         ing = ci.ingredients

#         total_qty = ci.quantity * persons
#         item_total = Decimal(total_qty) * Decimal(ing.price)

#         name = ing.name_en

#         if name in ingredient_summary:
#             ingredient_summary[name]["qty"] += total_qty
#             ingredient_summary[name]["total"] += item_total
#         else:
#             Ingredient_count += 1
#             ingredient_summary[name] = {
#                 "qty": total_qty,
#                 "unit": ing.unit,
#                 "price": ing.price,
#                 "total": item_total,
#             }

#         total_ingredient_cost += item_total

#     # =====================================================
#     # 3️⃣ COOKING ITEMS CALCULATION
#     # =====================================================
#     for item in cooking_items:

#         total_qty = 1   # because no quantity field
#         item_total = Decimal(item.cost)

#         name = item.name_en

#         cooking_summary[name] = {
#             "qty": total_qty,
#             "price": item.cost,
#             "total": item_total,
#         }
#         cooking_items_count += 1
#         total_cooking_cost += item_total



#     # -----------------------------------
#     # Create Beautiful PDF
#     # -----------------------------------
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="Invoice_{customer.id}.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         rightMargin=30,
#         leftMargin=30,
#         topMargin=30,
#         bottomMargin=30
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Custom Styles
#     styles.add(ParagraphStyle(
#         name='CenterTitle',
#         parent=styles['Title'],
#         alignment=1,
#         fontSize=20,
#         spaceAfter=20
#     ))

#     styles.add(ParagraphStyle(
#         name='SectionHeader',
#         parent=styles['Heading2'],
#         textColor=colors.HexColor("#0B5394"),
#         spaceBefore=15,
#         spaceAfter=10
#     ))

#     styles.add(ParagraphStyle(
#         name='TotalStyle',
#         parent=styles['Heading3'],
#         textColor=colors.HexColor("#1F4E79"),
#         spaceBefore=10
#     ))

#     # Title
#     elements.append(Paragraph("CATERING INVOICE", styles["CenterTitle"]))

#     # Invoice Info Box
#     invoice_data = [
#         ["Invoice No:", f"INV-{customer.id:04d}"],
#         ["Date:", datetime.now().strftime("%d-%m-%Y %I:%M %p")],
#     ]

#     invoice_table = Table(invoice_data, colWidths=[120, 300])
#     invoice_table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
#         ('BOX', (0, 0), (-1, -1), 1, colors.grey),
#         ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
#     ]))
#     elements.append(invoice_table)
#     elements.append(Spacer(1, 20))

#     # -----------------------------------
#     # Customer Details
#     # -----------------------------------
#     elements.append(Paragraph("Customer Details", styles["SectionHeader"]))

#     customer_data = [
#         ["Name", customer.name],
#         ["Phone", customer.phone],
#         ["Email", customer.email],
#         ["Address", customer.address],
#         ["No. of Persons", str(persons)],
#     ]

#     customer_table = Table(customer_data, colWidths=[120, 350])
#     customer_table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
#         ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
#         ('BOX', (0, 0), (-1, -1), 1, colors.grey),
#     ]))
#     elements.append(customer_table)
#     elements.append(Spacer(1, 20))

#     # -----------------------------------
#     # Selected Dishes
#     # -----------------------------------
#     elements.append(Paragraph("Selected Dishes", styles["SectionHeader"]))

#     dish_data = [["Dish Name"]]
#     for dish in dishes:
#         dish_data.append([dish.name_en])

#     dish_table = Table(dish_data, colWidths=[470])
#     dish_table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
#         ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
#     ]))
#     elements.append(dish_table)
#     elements.append(Spacer(1, 20))

#     # -----------------------------------
#     # Ingredient Summary
#     # -----------------------------------
#     elements.append(Paragraph("Ingredient Summary", styles["SectionHeader"]))

#     ingredient_data = [["Ingredient", "Qty", "Unit", "Price", "Total"]]

#     for name, data in ingredient_summary.items():
#         ingredient_data.append([
#             name,
#             str(data["qty"]),
#             data["unit"],
#             f"{data['price']:.2f}",
#             f"{data['total']:.2f}",
#         ])

#     ingredient_data.append([
#         "",
#         "",
#         "",
#         "Total:",
#         f"{total_ingredient_cost:.2f}"
#     ])

#     ingredient_table = Table(ingredient_data, colWidths=[110, 60, 70, 80, 90])
#     ingredient_table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#BDD7EE")),
#         ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
#         ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
#         ('BACKGROUND', (-2, -1), (-1, -1), colors.whitesmoke),
#     ]))
#     elements.append(ingredient_table)

#     elements.append(Paragraph(
#         f"Total Ingredient Items: {Ingredient_count}",
#         styles["TotalStyle"]
#     ))
#     elements.append(Spacer(1, 20))

#     # -----------------------------------
#     # Cooking Items
#     # -----------------------------------
#     elements.append(Paragraph("Cooking Items", styles["SectionHeader"]))

#     cooking_data = [["Item", "Qty", "Total"]]

#     for name, data in cooking_summary.items():
#         cooking_data.append([
#             name,
#             str(data["qty"]),
#             f"{data['total']:.2f}"
#         ])

#     cooking_data.append([
#         "",
#         "Total:",
#         f"{total_cooking_cost:.2f}"
#     ])

#     cooking_table = Table(cooking_data, colWidths=[200, 80, 120])
#     cooking_table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#BDD7EE")),
#         ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
#         ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
#         ('BACKGROUND', (-2, -1), (-1, -1), colors.whitesmoke),
#     ]))
#     elements.append(cooking_table)

#     elements.append(Paragraph(
#         f"Total Cooking Items: {cooking_items_count}",
#         styles["TotalStyle"]
#     ))
#     elements.append(Spacer(1, 25))

#     # -----------------------------------
#     # GRAND TOTAL
#     # -----------------------------------
#     grand_total = total_ingredient_cost + total_cooking_cost

#     grand_table = Table(
#         [["GRAND TOTAL", f"{grand_total:.2f}"]],
#         colWidths=[300, 100]
#     )

#     grand_table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#0B5394")),
#         ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
#         ('ALIGN', (1, 0), (1, 0), 'CENTER'),
#         ('FONTSIZE', (0, 0), (-1, -1), 14),
#         ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
#     ]))

#     elements.append(grand_table)
#     elements.append(Spacer(1, 40))

#     elements.append(Paragraph(
#         "Thank you for choosing our catering service!",
#         styles["Normal"]
#     ))

#     doc.build(elements)
#     return response

def export_customer_pdf(request, customer_id):

    customer = get_object_or_404(Customer, id=customer_id)
    owner = request.user

    dishes = customer.dishes.all()
    cooking_items = customer.cooking_items.all()
    customer_ingredients = customer.customeringredient_set.all()

    persons = customer.num_person or 1

    Ingredient_count = 0
    cooking_items_count = 0
    ingredient_summary = {}
    cooking_summary = {}

    total_ingredient_cost = Decimal("0.00")
    total_cooking_cost = Decimal("0.00")

    # ---------------- Ingredient Calculation ----------------
    for dish in dishes:
        dish_ingredients = DishIngredient.objects.filter(dish=dish)

        for di in dish_ingredients:
            total_qty = di.quantity * persons
            total_price = Decimal(total_qty) * Decimal(di.price)

            name = di.ingredient.name_en

            if name in ingredient_summary:
                ingredient_summary[name]["qty"] += total_qty
                ingredient_summary[name]["total"] += total_price
            else:
                Ingredient_count += 1
                ingredient_summary[name] = {
                    "qty": total_qty,
                    "unit": di.unit,
                    "price": di.price,
                    "total": total_price,
                }

            total_ingredient_cost += total_price

    for ci in customer_ingredients:
        ing = ci.ingredients
        total_qty = ci.quantity * persons
        item_total = Decimal(total_qty) * Decimal(ing.price)

        name = ing.name_en

        if name in ingredient_summary:
            ingredient_summary[name]["qty"] += total_qty
            ingredient_summary[name]["total"] += item_total
        else:
            Ingredient_count += 1
            ingredient_summary[name] = {
                "qty": total_qty,
                "unit": ing.unit,
                "price": ing.price,
                "total": item_total,
            }

        total_ingredient_cost += item_total

    # ---------------- Cooking Items ----------------
    for item in cooking_items:
        item_total = Decimal(item.cost)
        name = item.name_en

        cooking_summary[name] = {
            "qty": 1,
            "price": item.cost,
            "total": item_total,
        }

        cooking_items_count += 1
        total_cooking_cost += item_total

    grand_total = total_ingredient_cost + total_cooking_cost

    # ---------------- Create PDF ----------------
    response = HttpResponse(content_type='application/pdf')
    # response['Content-Disposition'] = f'attachment; filename="Invoice_{customer.id}.pdf"'
    response['Content-Disposition'] = f'inline; filename="Invoice_{customer.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name='CenterTitle',
        parent=styles['Title'],
        alignment=1,
        fontSize=22,
        spaceAfter=20
    ))

    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading2'],
        textColor=colors.HexColor("#1F4E79"),
        spaceBefore=15,
        spaceAfter=8
    ))

    # ---------------- Logo ----------------
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/CMS_logo.jpg')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=120, height=60)
        elements.append(logo)

    elements.append(Spacer(1, 10))

    # ---------------- Title ----------------
    elements.append(Paragraph("CATERING INVOICE", styles["CenterTitle"]))

    # ---------------- Invoice Info ----------------
    invoice_data = [
        ["Invoice No:", f"INV-{customer.id:04d}"],
        ["Date:", datetime.now().strftime("%d-%m-%Y %I:%M %p")],
    ]

    invoice_table = Table(invoice_data, colWidths=[120, 300])
    invoice_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(invoice_table)
    elements.append(Spacer(1, 20))

    # ---------------- Proparator Details ----------------
    elements.append(Paragraph("Proparator Details", styles["SectionHeader"]))

    owner_data = [
        ["Name", owner.username],
        ["Phone", owner.phone_number],
        ["Email", owner.email],
        ["Address", owner.address],
    ]

    owner_table = Table(owner_data, colWidths=[120, 350])
    owner_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(owner_table)
    elements.append(Spacer(1, 20))

    # ---------------- Customer Details ----------------
    elements.append(Paragraph("Customer Details", styles["SectionHeader"]))

    customer_data = [
        ["Name", customer.name],
        ["Phone", customer.phone],
        ["Email", customer.email],
        ["Address", customer.address],
        ["Reason", customer.reason],
        ["Persons", str(persons)],
    ]

    customer_table = Table(customer_data, colWidths=[120, 350])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(customer_table)
    elements.append(Spacer(1, 20))

    # ---------------- Ingredient Table ----------------
    elements.append(Paragraph("Ingredient Summary", styles["SectionHeader"]))

    ingredient_data = [["Ingredient", "Qty", "Unit", "Price", "Total"]]

    for name, data in ingredient_summary.items():
        ingredient_data.append([
            name,
            str(data["qty"]),
            data["unit"],
            f"{data['price']:.2f}",
            f"{data['total']:.2f}",
        ])

    ingredient_data.append(["", "", "", "Total:", f"{total_ingredient_cost:.2f}"])
    ingredient_data.append(["", "", "", "Total Count:", f"{Ingredient_count:.2f}"])
    ingredient_table = Table(ingredient_data, colWidths=[110, 60, 70, 80, 90])
    ingredient_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#BDD7EE")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
    ]))

    elements.append(ingredient_table)
    elements.append(Spacer(1, 20))

    

    # ---------------- Cooking Table ----------------
    elements.append(Paragraph("Cooking Items", styles["SectionHeader"]))

    cooking_data = [["Item", "Qty", "Total"]]

    for name, data in cooking_summary.items():
        cooking_data.append([
            name,
            str(data["qty"]),
            f"{data['total']:.2f}"
        ])

    cooking_data.append(["", "Total:", f"{total_cooking_cost:.2f}"])
    cooking_data.append(["", "Total Count:", f"{cooking_items_count:.2f}"])
    cooking_table = Table(cooking_data, colWidths=[200, 80, 120])
    cooking_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#BDD7EE")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
    ]))

    elements.append(cooking_table)
    elements.append(Spacer(1, 30))

    # ---------------- Grand Total Box ----------------
    grand_table = Table(
        [["GRAND TOTAL", f"{grand_total:.2f}"]],
        colWidths=[300, 100]
    )

    grand_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#1F4E79")),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
    ]))

    elements.append(grand_table)
    elements.append(Spacer(1, 50))

    # ---------------- Signature ----------------
    elements.append(Paragraph("Authorized Signature", styles["Normal"]))
    elements.append(Spacer(1, 10))
    elements.append(Table([["_________________________"]], colWidths=[200]))

    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Thank you for choosing our catering service!", styles["Normal"]))

    doc.build(
        elements,
        onFirstPage=lambda c, d: (add_watermark(c, d), add_page_number(c, d)),
        onLaterPages=lambda c, d: (add_watermark(c, d), add_page_number(c, d))
    )

    return response


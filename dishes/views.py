from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
from django.core.files.base import ContentFile
import requests

from .models import Dish, DishIngredient
from .forms import DishForm, DishIngredientFormSet, DishIngredientForm
from ingredients.models import Ingredient

from django.forms import inlineformset_factory
from dishes.models import Dish

# =========================
# LIST
# =========================

from django.db.models import Q
from django.contrib.auth.decorators import login_required

@login_required
def dish_list(request):
    user_lang = getattr(request.user, "preferred_language", "en")
    query = request.GET.get("q")

    dishes = Dish.objects.prefetch_related(
        "dishingredient_set__ingredient"
    ).order_by("dish_id")

    # ✅ SEARCH FILTER
    if query:
        dishes = dishes.filter(
            Q(dish_id__icontains=query) |
            Q(name_en__icontains=query) |
            Q(**{f"name_{user_lang}__icontains": query}) |
            Q(dishingredient__ingredient__name_en__icontains=query) |
            Q(**{f"dishingredient__ingredient__name_{user_lang}__icontains": query})
        ).distinct()

    for dish in dishes:
        dish.display_name = getattr(dish, f"name_{user_lang}", None) or dish.name_en
        dish.display_preparation = getattr(dish, f"preparation_{user_lang}", None) or dish.preparation_en

        dish.display_ingredients = []

        total_cost = 0

        for di in dish.dishingredient_set.all():
            ing_name = getattr(di.ingredient, f"name_{user_lang}", None) or di.ingredient.name_en

            dish.display_ingredients.append({
                "name": ing_name,
                "quantity": di.quantity,
                "unit": di.unit,
                "price": di.price
            })

            total_cost += di.quantity * di.price

        dish.total_cost = total_cost

    return render(request, "dishes/dish_list.html", {
        "dishes": dishes,
        "query": query
    })



# =========================
# ADD
# =========================
# @login_required
# def add_dish(request):
#     ingredient_list = Ingredient.objects.all()

#     if request.method == "POST":
#         form = DishForm(request.POST, request.FILES)
#         formset = DishIngredientFormSet(request.POST, request.FILES)

#         if form.is_valid() and formset.is_valid():

#             dish = form.save(commit=False)

#             if not dish.dish_id:
#                 last = Dish.objects.order_by("-id").first()
#                 number = int(last.dish_id[1:]) + 1 if last else 1
#                 dish.dish_id = f"D{number:03d}"

#             dish.save()

#             formset.instance = dish
#             formset.save()

#             return redirect("dish_list")   # ✅ WILL NOW WORK

#         else:
#             print(form.errors)
#             print(formset.errors)

#     else:
#         form = DishForm()
#         formset = DishIngredientFormSet()

#     return render(request, "dishes/add_dish.html", {
#         "form": form,
#         "formset": formset,
#         "ingredient_list": ingredient_list,
#         "title": "Add Dish"
#     })

@login_required
def add_dish(request):
    ingredient_list = Ingredient.objects.all()

    if request.method == "POST":
        form = DishForm(request.POST, request.FILES)
        formset = DishIngredientFormSet(request.POST)

        if form.is_valid():
            dish = form.save(commit=False)

            if not dish.dish_id:
                last = Dish.objects.order_by("-id").first()
                number = int(last.dish_id[1:]) + 1 if last else 1
                dish.dish_id = f"D{number:03d}"

            dish.save()

            formset = DishIngredientFormSet(request.POST, instance=dish)

            if formset.is_valid():
                formset.save()
                messages.success(request, "Dish added successfully!")
                return redirect("dish_list")
            else:
                print("FORMSET ERRORS:", formset.errors)

        else:
            print("FORM ERRORS:", form.errors)

    else:
        form = DishForm()
        formset = DishIngredientFormSet()

    return render(request, "dishes/add_dish.html", {
        "form": form,
        "formset": formset,
        "ingredient_list": ingredient_list,
        "title": "Add Dish"
    })


# =========================
# EDIT
# =========================
# @login_required
# def edit_dish(request, pk):
#     dish = get_object_or_404(Dish, pk=pk)
#     ingredient_list = Ingredient.objects.all()

#     if request.method == "POST":
#         form = DishForm(request.POST, request.FILES, instance=dish)
#         formset = DishIngredientFormSet(request.POST, request.FILES, instance=dish)

#         if form.is_valid() and formset.is_valid():
#             form.save()
#             formset.save()
#             return redirect("dish_list")

#         else:
#             print(form.errors)
#             print(formset.errors)

#     else:
#         form = DishForm(instance=dish)
#         formset = DishIngredientFormSet(instance=dish)

#     return render(request, "dishes/add_dish.html", {
#         "form": form,
#         "formset": formset,
#         "ingredient_list": ingredient_list,
#         "title": "Edit Dish"
#     })

@login_required
def edit_dish(request, pk):
    dish = get_object_or_404(Dish, pk=pk)
    ingredient_list = Ingredient.objects.all()

    if request.method == "POST":
        form = DishForm(request.POST, request.FILES, instance=dish)
        formset = DishIngredientFormSet(request.POST, instance=dish)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Dish updated successfully!")
            return redirect("dish_list")
        else:
            print("FORM ERRORS:", form.errors)
            print("FORMSET ERRORS:", formset.errors)

    else:
        form = DishForm(instance=dish)
        formset = DishIngredientFormSet(instance=dish)

    return render(request, "dishes/add_dish.html", {
        "form": form,
        "formset": formset,
        "ingredient_list": ingredient_list,
        "title": "Edit Dish"
    })


# =========================
# DELETE
# =========================
@login_required
def delete_dish(request, pk):
    dish = get_object_or_404(Dish, pk=pk)
    dish.delete()
    messages.success(request, "Dish deleted successfully.")
    return redirect("dish_list")


# =========================
# EXPORT
# =========================
@login_required
def export_dishes(request):
    dishes = Dish.objects.prefetch_related("dishingredient_set__ingredient")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dishes"

    ws.append([
        "dish_id",
        "name_en", "name_te", "name_ta", "name_hi", "name_ka",
        "prep_en", "prep_te", "prep_ta", "prep_hi", "prep_ka",
        "image_url",
        "ingredient",
        "quantity",
        "unit",
        "price",
        "row_total"
    ])

    for dish in dishes:
        for di in dish.dishingredient_set.all():
            ws.append([
                dish.dish_id,
                dish.name_en,
                dish.name_te,
                dish.name_ta,
                dish.name_hi,
                dish.name_ka,
                dish.preparation_en,
                dish.preparation_te,
                dish.preparation_ta,
                dish.preparation_hi,
                dish.preparation_ka,
                "", 
                di.ingredient.name_en,
                di.quantity,
                di.unit,
                di.price,
                di.quantity * di.price
            ])
            if dish.image:
                try:
                    img = OpenpyxlImage(dish.image.path)
                    img.width = 80
                    img.height = 80
                    ws.add_image(img, f"L{ws.max_row}")  # Adjust column if needed
                except:
                    pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=full_dishes_export.xlsx'
    return response


# =========================
# IMPORT
# =========================
# @login_required
# def import_dishes(request):
#     if request.method == "POST" and request.FILES.get("excel_file"):

#         wb = load_workbook(request.FILES["excel_file"])
#         ws = wb.active

#         for row in ws.iter_rows(min_row=2, values_only=True):

#             (
#                 dish_id,
#                 name_en, name_te, name_ta, name_hi, name_ka,
#                 prep_en, prep_te, prep_ta, prep_hi, prep_ka,
#                 image_url,
#                 ingredient_name,
#                 quantity,
#                 unit,
#                 price,
#                 row_total
#             ) = row

#             dish, created = Dish.objects.get_or_create(
#                 dish_id=dish_id,
#                 defaults={
#                     "name_en": name_en,
#                     "name_te": name_te,
#                     "name_ta": name_ta,
#                     "name_hi": name_hi,
#                     "name_ka": name_ka,
#                     "preparation_en": prep_en,
#                     "preparation_te": prep_te,
#                     "preparation_ta": prep_ta,
#                     "preparation_hi": prep_hi,
#                     "preparation_ka": prep_ka,
#                 }
#             )

#             # 🔥 Image Download
#             if image_url:
#                 try:
#                     response = requests.get(image_url)
#                     if response.status_code == 200:
#                         dish.image.save(
#                             f"{dish_id}.jpg",
#                             ContentFile(response.content),
#                             save=True
#                         )
#                 except:
#                     pass

#             try:
#                 ingredient = Ingredient.objects.get(name_en=ingredient_name)

#                 DishIngredient.objects.update_or_create(
#                     dish=dish,
#                     ingredient=ingredient,
#                     defaults={
#                         "quantity": quantity or 0,
#                         "unit": unit or "N/A",
#                         "price": price or 0
#                     }
#                 )

#             except Ingredient.DoesNotExist:
#                 continue

#         return redirect("dish_list")

#     return render(request, "dishes/import_dishes.html")

@login_required
def import_dishes(request):
    if request.method == "POST" and request.FILES.get("excel_file"):

        wb = load_workbook(request.FILES["excel_file"])
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]

            (
                dish_id,
                name_en, name_te, name_ta, name_hi, name_ka,
                prep_en, prep_te, prep_ta, prep_hi, prep_ka,
                image_url,
                ingredient_name,
                quantity,
                unit,
                price,
                row_total
            ) = values[:17]

            dish, created = Dish.objects.get_or_create(
                dish_id=dish_id,
                defaults={
                    "name_en": name_en,
                    "name_te": name_te,
                    "name_ta": name_ta,
                    "name_hi": name_hi,
                    "name_ka": name_ka,
                    "preparation_en": prep_en,
                    "preparation_te": prep_te,
                    "preparation_ta": prep_ta,
                    "preparation_hi": prep_hi,
                    "preparation_ka": prep_ka,
                }
            )

            # 🔥 Download image from URL
            if image_url:
                try:
                    response = requests.get(image_url)
                    if response.status_code == 200:
                        dish.image.save(
                            f"{dish_id}.jpg",
                            ContentFile(response.content),
                            save=True
                        )
                except:
                    pass

            # 🔥 Handle Embedded Excel Images
            for img in ws._images:
                try:
                    row_number = img.anchor._from.row + 1
                    excel_dish_id = ws.cell(row=row_number, column=1).value

                    if excel_dish_id == dish_id:
                        from PIL import Image as PILImage
                        pil_img = PILImage.open(img.ref)
                        pil_img.thumbnail((300,300))

                        with BytesIO() as img_bytes:
                            pil_img.save(img_bytes, format="PNG")
                            img_bytes.seek(0)

                            dish.image.save(
                                f"{dish_id}.png",
                                ContentFile(img_bytes.read()),
                                save=True
                            )
                except:
                    pass

            try:
                ingredient = Ingredient.objects.get(name_en=ingredient_name)

                DishIngredient.objects.update_or_create(
                    dish=dish,
                    ingredient=ingredient,
                    defaults={
                        "quantity": quantity or 0,
                        "unit": unit or "pcs",
                        "price": price or 0
                    }
                )
            except Ingredient.DoesNotExist:
                continue

        return redirect("dish_list")

    return render(request, "dishes/import_dishes.html")

    
@login_required
def dashboard_dish_count(request):
    ingredient_count = Ingredient.objects.count()
    dish_count = Dish.objects.count()
    
    return render(request, "dashboard.html", {
        "dish_count": dish_count,
    })

@login_required
def get_ingredient_details(request):
    ingredient_id = request.GET.get("ingredient_id")
    try:
        ingredient = Ingredient.objects.get(id=ingredient_id)
        return JsonResponse({
            "quantity": ingredient.quantity,
            "unit": ingredient.unit,
            "price": ingredient.price,
        })
    except Ingredient.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

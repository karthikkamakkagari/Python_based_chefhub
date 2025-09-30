from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.files.base import ContentFile
import pandas as pd
import requests
from .models import Ingredient
from .forms import IngredientForm
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
from PIL import Image as PILImage

@login_required
def ingredient_list(request):
    ingredients = Ingredient.objects.all()
    return render(request, 'ingredients/ingredient_list.html', {'items': ingredients, 'title':'Ingredients'})

@login_required
def add_edit_ingredient(request, pk=None):
    if pk:
        instance = get_object_or_404(Ingredient, pk=pk)
        title = "Edit Ingredient"
    else:
        instance = None
        title = "Add Ingredient"

    if request.method == 'POST':
        form = IngredientForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            ingredient = form.save(commit=False)
            # Handle image URL download
            image_url = form.cleaned_data.get('image_url')
            if image_url and not ingredient.image:
                try:
                    response = requests.get(image_url)
                    if response.status_code == 200:
                        name = image_url.split('/')[-1]
                        ingredient.image.save(name, ContentFile(response.content), save=False)
                except:
                    pass
            ingredient.save()
            return redirect('ingredient_list')
    else:
        form = IngredientForm(instance=instance)

    return render(request, 'ingredients/add_edit_ingredient.html', {'form':form, 'title':title})

@login_required
def delete_ingredient(request, pk):
    ingredient = get_object_or_404(Ingredient, pk=pk)
    if request.method == "POST":
        ingredient.delete()
        return redirect('ingredient_list')
    return render(request, 'ingredients/delete_ingredient.html', {'ingredient':ingredient})

@login_required
# def export_ingredients(request):
#     ingredients = Ingredient.objects.all().values()
#     df = pd.DataFrame(list(ingredients))
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=ingredients.xlsx'
#     df.to_excel(response, index=False)
#     return response
def export_ingredients(request):
    ingredients = Ingredient.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingredients"

    # Header
    headers = ['ingredient_id','name_en','name_te','name_ta','name_hi','name_ka',
               'quantity','unit','price','image_url','image']
    ws.append(headers)

    for ing in ingredients:
        row = [
            ing.ingredient_id,
            ing.name_en,
            ing.name_te,
            ing.name_ta,
            ing.name_hi,
            ing.name_ka,
            ing.quantity,
            ing.unit,
            ing.price,
            ing.image_url,
            ''  # Placeholder for image
        ]
        ws.append(row)
        
        # Insert image if exists
        if ing.image:
            img_path = ing.image.path
            img = OpenpyxlImage(img_path)
            # Resize image for Excel (optional)
            img.width = 80
            img.height = 80
            img_cell = f'K{ws.max_row}'  # Column K is for image
            ws.add_image(img, img_cell)

    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ingredients_with_images.xlsx'
    return response

@login_required
def import_ingredients(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(filename=excel_file)
        ws = wb.active

        # Map column headers
        headers = [cell.value for cell in ws[1]]
        col_index = {header:i+1 for i, header in enumerate(headers)}  # 1-based index

        # Step 1: Create/update ingredients (fields only)
        for row in ws.iter_rows(min_row=2):
            data = {header: row[col_index[header]-1].value if header in col_index else None
                    for header in headers}

            ingredient, created = Ingredient.objects.update_or_create(
                ingredient_id=data.get('ingredient_id'),
                defaults={
                    'name_en': data.get('name_en',''),
                    'name_te': data.get('name_te',''),
                    'name_ta': data.get('name_ta',''),
                    'name_hi': data.get('name_hi',''),
                    'name_ka': data.get('name_ka',''),
                    'quantity': data.get('quantity',0),
                    'unit': data.get('unit','pcs'),
                    'price': data.get('price',0),
                    'image_url': data.get('image_url',''),
                }
            )

            # Step 1b: If image_url exists, download image
            image_url = data.get('image_url')
            if image_url:
                try:
                    response = requests.get(image_url)
                    if response.status_code == 200:
                        name = image_url.split('/')[-1]
                        ingredient.image.save(name, ContentFile(response.content), save=True)
                except:
                    pass

        # Step 2: Optional – handle embedded images in Excel (only if you use embedded)
        for img in ws._images:
            try:
                # Determine row in Excel
                if hasattr(img.anchor, 'tl'):  # newer openpyxl
                    row_number = img.anchor.tl.row + 1
                else:  # fallback
                    row_number = img.anchor._from.row + 1

                # Assume ingredient_id is in column A
                ingredient_id_cell = ws.cell(row=row_number, column=1).value
                if not ingredient_id_cell:
                    continue

                ingredient = Ingredient.objects.get(ingredient_id=ingredient_id_cell)

                # Save image
                pil_img = PILImage.open(img.ref)
                pil_img.thumbnail((200,200))
                with BytesIO() as img_bytes:
                    pil_img.save(img_bytes, format='PNG')
                    img_bytes.seek(0)
                    ingredient.image.save(f"{ingredient.ingredient_id}_{ingredient.name_en}.png", ContentFile(img_bytes.read()), save=True)

            except Exception as e:
                print(f"Error processing image in row {row_number}: {e}")

        return redirect('ingredient_list')

    return render(request, 'ingredients/import_ingredients.html')


# --- New Dashboard View ---
@login_required
def dashboard_ingredient_count(request):
    ingredient_count = Ingredient.objects.count()
    context = {
        'ingredient_count': ingredient_count
    }
    return render(request, 'dashboard.html', context)
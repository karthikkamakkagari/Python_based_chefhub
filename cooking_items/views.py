from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from .models import CookingItem
from .forms import CookingItemForm
from django.conf import settings
from django.core.files import File
import os
from django.core.files.base import ContentFile
import uuid



# =========================
# LIST + SEARCH
# =========================
@login_required
def cooking_item_list(request):
    user_lang = getattr(request.user, "preferred_language", "en")
    query = request.GET.get("q")

    items = CookingItem.objects.all().order_by("item_id")

    # ✅ SEARCH
    if query:
        items = items.filter(
            Q(item_id__icontains=query) |
            Q(name_en__icontains=query) |
            Q(**{f"name_{user_lang}__icontains": query})
        )

    # ✅ Display language based on profile
    for item in items:
        item.display_name = getattr(item, f"name_{user_lang}", None) or item.name_en
        item.display_summary = getattr(item, f"summary_{user_lang}", None) or item.summary_en

    return render(request, "cooking_items/cooking_item_list.html", {
        "items": items,
        "query": query,
        "total_items": items.count()
    })

@login_required
def add_cooking_item(request):

    if request.method == "POST":
        form = CookingItemForm(request.POST, request.FILES)

        if form.is_valid():
            item = form.save(commit=False)

            # 🔥 AUTO NUMERIC ID
            if not item.item_id:
                last = CookingItem.objects.order_by("-id").first()
                number = int(last.item_id) + 1 if last else 1
                item.item_id = str(number)

            item.save()
            messages.success(request, "Cooking Item added successfully!")
            return redirect("cooking_item_list")

    else:
        form = CookingItemForm()

    return render(request, "cooking_items/add_cooking_item.html", {
        "form": form,
        "title": "Add Cooking Item"
    })

@login_required
def edit_cooking_item(request, pk):
    item = get_object_or_404(CookingItem, pk=pk)

    if request.method == "POST":
        form = CookingItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Cooking Item updated successfully!")
            return redirect("cooking_item_list")
    else:
        form = CookingItemForm(instance=item)

    return render(request, "cooking_items/add_cooking_item.html", {
        "form": form,
        "title": "Edit Cooking Item"
    })

@login_required
def delete_cooking_item(request, pk):
    item = get_object_or_404(CookingItem, pk=pk)
    item.delete()
    messages.success(request, "Cooking Item deleted successfully.")
    return redirect("cooking_item_list")


@login_required
def export_cooking_items(request):
    items = CookingItem.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cooking Items"

    # ✅ Header (Image column added at end)
    ws.append([
        "Item ID",
        "Name (English)",
        "Name (Telugu)",
        "Name (Tamil)",
        "Name (Hindi)",
        "Name (Kannada)",
        "Summary (English)",
        "Summary (Telugu)",
        "Summary (Tamil)",
        "Summary (Hindi)",
        "Summary (Kannada)",
        "Quantity",
        "Cost",
        "Image"
    ])

    row_num = 2

    for item in items:
        ws.cell(row=row_num, column=1, value=item.item_id)
        ws.cell(row=row_num, column=2, value=item.name_en)
        ws.cell(row=row_num, column=3, value=item.name_te)
        ws.cell(row=row_num, column=4, value=item.name_ta)
        ws.cell(row=row_num, column=5, value=item.name_hi)
        ws.cell(row=row_num, column=6, value=item.name_ka)
        ws.cell(row=row_num, column=7, value=item.summary_en)
        ws.cell(row=row_num, column=8, value=item.summary_te)
        ws.cell(row=row_num, column=9, value=item.summary_ta)
        ws.cell(row=row_num, column=10, value=item.summary_hi)
        ws.cell(row=row_num, column=11, value=item.summary_ka)
        ws.cell(row=row_num, column=12, value=item.quantity)
        ws.cell(row=row_num, column=13, value=item.cost)

        # ✅ ADD IMAGE TO COLUMN 14
        if item.image and os.path.exists(item.image.path):
            img = ExcelImage(item.image.path)
            img.width = 80
            img.height = 80
            ws.add_image(img, f"N{row_num}")  # Column N = 14

        row_num += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cooking_items.xlsx'

    wb.save(response)
    return response

@login_required
def import_cooking_items(request):

    if request.method == "POST":
        print("IMPORT VIEW CALLED")
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return redirect('cooking_item_list')

        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        images = ws._images

        # 🔥 Get real last row by checking column 1 manually
        row_num = 2

        while True:
            first_cell = ws.cell(row=row_num, column=1).value

            if first_cell is None:
                break

            item = CookingItem.objects.create(
                item_id=ws.cell(row=row_num, column=1).value,
                name_en=ws.cell(row=row_num, column=2).value,
                name_te=ws.cell(row=row_num, column=3).value,
                name_ta=ws.cell(row=row_num, column=4).value,
                name_hi=ws.cell(row=row_num, column=5).value,
                name_ka=ws.cell(row=row_num, column=6).value,
                summary_en=ws.cell(row=row_num, column=7).value,
                summary_te=ws.cell(row=row_num, column=8).value,
                summary_ta=ws.cell(row=row_num, column=9).value,
                summary_hi=ws.cell(row=row_num, column=10).value,
                summary_ka=ws.cell(row=row_num, column=11).value,
                quantity=ws.cell(row=row_num, column=12).value,
                cost=ws.cell(row=row_num, column=13).value,
            )

            # ✅ Attach image by row
            for img in images:
                if img.anchor._from.row + 1 == row_num:
                    image_bytes = img._data()
                    image_file = ContentFile(image_bytes)
                    filename = f"{uuid.uuid4()}.png"
                    item.image.save(filename, image_file, save=True)

            row_num += 1
        print(ws.max_row)
        print(ws.cell(row=2, column=1).value)
        return redirect('cooking_item_list')

    return render(request, "cooking_items/cooking_item_import.html")

@login_required
def dashboard_cooking_item_count(request):
    cooking_item_count = CookingItem.objects.count()
    context = {

        'cooking_item_count': cooking_item_count,
    }
    return render(request, 'dashboard.html', context)